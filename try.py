#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
機器帳號爬蟲 - 從總後台 API 抓取機器的平台帳號
輸出：Excel 檔案（只顯示 WM/AB/MT/T9/SA 五個平台）
"""

import json
import requests
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ==================== 配置區 ====================
API_BASE_URL = "https://wpapi.ldjzmr.top/master"
BEARER_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczovL3dwYXBpLmxkanptci50b3AvbWFzdGVyL2xvZ2luIiwiaWF0IjoxNzcwNDI5NjIxLCJleHAiOjE4MDE5NjU2MjEsIm5iZiI6MTc3MDQyOTYyMSwianRpIjoicXpGSUx5c296eHZPczhyTSIsInN1YiI6IjExIiwicHJ2IjoiMTg4ODk5NDM5MDUwZTVmMzc0MDliMThjYzZhNDk1NjkyMmE3YWIxYiJ9.FJwCCTCn6CmghjL6gCTxyVDwa9-UZH25GiHT_JrIhYg"

# 輸出檔案路徑（py 檔所在資料夾）
OUTPUT_DIR = Path(__file__).parent

# ==================== 工具函數 ====================
def log(msg: str):
    """帶時間戳記的 log"""
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}")


def fetch_machines(page: int = 1, page_size: int = 100):
    """
    從 API 獲取機器列表
    
    Args:
        page: 頁碼
        page_size: 每頁筆數
    
    Returns:
        dict: API 回傳的 JSON 資料
    """
    url = f"{API_BASE_URL}/machine"
    headers = {
        "Authorization": f"Bearer {BEARER_TOKEN}",
        "Content-Type": "application/json"
    }
    params = {
        "pagenum": page,
        "pagesize": page_size
    }
    
    try:
        response = requests.get(url, headers=headers, params=params, timeout=30)
        response.raise_for_status()
        return response.json()
        
    except requests.exceptions.RequestException as e:
        log(f"❌ API 請求失敗: {e}")
        return None


def parse_machine_data(api_response):
    """
    解析 API 回傳的機器資料
    
    Args:
        api_response: API 回傳的 JSON
    
    Returns:
        list: 整理後的機器資料列表
    """
    if not api_response:
        return []
    
    machines = []
    items = []
    
    # 找出資料陣列（data.data）
    if isinstance(api_response, dict) and "data" in api_response:
        data = api_response["data"]
        if isinstance(data, dict) and "data" in data:
            items = data["data"]
    
    for item in items:
        try:
            # 機器基本資料（外層）
            machine_id = str(item.get("id", ""))
            machine_name = item.get("name", "")
            
            # 商戶名稱（在 brand 物件裡）
            brand = item.get("brand", {})
            brand_name = brand.get("name", "") if brand else ""
            
            # 平台帳號（在 user 物件裡）
            user = item.get("user", {})
            if user:
                wm_id = user.get("WM_id") or ""
                ab_id = user.get("AB_id") or ""
                mt_id = user.get("MT_id") or ""
                t9_id = user.get("T9_id") or ""
                sa_id = user.get("SA_id") or ""
            else:
                wm_id = ab_id = mt_id = t9_id = sa_id = ""
            
            machine = {
                "機器ID": machine_id,
                "商戶名稱": brand_name,
                "機器名稱": machine_name,
                "WM帳號": wm_id,
                "AB帳號": ab_id,
                "MT帳號": mt_id,
                "T9帳號": t9_id,
                "SA帳號": sa_id,
            }
            machines.append(machine)
            
        except Exception as e:
            continue
    
    return machines


def create_excel(data: list, filename: str = "機器帳號列表.xlsx"):
    """
    創建 Excel 檔案
    
    Args:
        data: list of dict
        filename: 輸出檔名
    
    Returns:
        Path: Excel 檔案路徑
    """
    log(f"📊 開始建立 Excel：{filename}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "機器帳號"
    
    # 設定表頭（加上商戶名稱）
    headers = ["機器ID", "商戶名稱", "機器名稱", "WM帳號", "AB帳號", "MT帳號", "T9帳號", "SA帳號"]
    
    # 寫入表頭（加粗、置中、背景色）
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 寫入資料
    for row_num, item in enumerate(data, 2):
        ws.cell(row=row_num, column=1, value=item.get("機器ID", ""))
        ws.cell(row=row_num, column=2, value=item.get("商戶名稱", ""))
        ws.cell(row=row_num, column=3, value=item.get("機器名稱", ""))
        ws.cell(row=row_num, column=4, value=item.get("WM帳號", ""))
        ws.cell(row=row_num, column=5, value=item.get("AB帳號", ""))
        ws.cell(row=row_num, column=6, value=item.get("MT帳號", ""))
        ws.cell(row=row_num, column=7, value=item.get("T9帳號", ""))
        ws.cell(row=row_num, column=8, value=item.get("SA帳號", ""))
    
    # 自動調整欄寬
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # 凍結首列
    ws.freeze_panes = "A2"
    
    # 儲存
    output_path = OUTPUT_DIR / filename
    wb.save(output_path)
    log(f"✅ Excel 已儲存：{output_path}")
    
    return output_path


# ==================== 主程式 ====================
def main():
    """主程式入口"""
    log("=" * 60)
    log("機器帳號爬蟲程式 - 只抓 WM/AB/MT/T9/SA 五個平台")
    log("=" * 60)
    
    all_machines = []
    page = 1
    page_size = 1000
    
    while True:
        log(f"\n📄 正在抓取第 {page} 頁...")
        
        api_response = fetch_machines(page=page, page_size=page_size)
        
        if not api_response:
            log("❌ 無法獲取資料，停止抓取")
            break
        
        machine_data = parse_machine_data(api_response)
        
        if not machine_data:
            log(f"⚠️ 第 {page} 頁沒有資料，抓取完成")
            break
        
        log(f"✅ 第 {page} 頁：解析到 {len(machine_data)} 筆資料")
        all_machines.extend(machine_data)
        
        if len(machine_data) < page_size:
            log(f"📌 已到達最後一頁（共 {page} 頁）")
            break
        
        page += 1
    
    if not all_machines:
        log("⚠️ 沒有解析到任何機器資料")
        return
    
    log(f"\n✅ 總共抓取 {len(all_machines)} 筆機器資料（共 {page} 頁）")
    
    # 統計有帳號的數量
    wm_count = sum(1 for m in all_machines if m.get("WM帳號"))
    ab_count = sum(1 for m in all_machines if m.get("AB帳號"))
    mt_count = sum(1 for m in all_machines if m.get("MT帳號"))
    t9_count = sum(1 for m in all_machines if m.get("T9帳號"))
    sa_count = sum(1 for m in all_machines if m.get("SA帳號"))
    
    log(f"\n📊 帳號統計：")
    log(f"   - WM帳號：{wm_count} 筆")
    log(f"   - AB帳號：{ab_count} 筆")
    log(f"   - MT帳號：{mt_count} 筆")
    log(f"   - T9帳號：{t9_count} 筆")
    log(f"   - SA帳號：{sa_count} 筆")
    
    # 生成 Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"機器帳號列表_{timestamp}.xlsx"
    excel_path = create_excel(all_machines, filename)
    
    log("=" * 60)
    log(f"🎉 完成！")
    log(f"📁 檔案位置：{excel_path}")
    log(f"📊 共 {len(all_machines)} 筆機器資料")
    log("=" * 60)


if __name__ == "__main__":
    main()