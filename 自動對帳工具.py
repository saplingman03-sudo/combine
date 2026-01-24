import requests
import pandas as pd
import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
import json
import os

# ============================================================
# 王牌財務工具 V2.1
# 功能：
# 1) 依時間區間抓取帳務流水 (banknote_log)
# 2) 依店家(brand)彙總：開分/投鈔/洗分、當月累計營業額、前日累計額、今日變化
# 3) 支援「特殊結算店家」：各店家的月初起算點可不同(01號/月底 + 指定小時)
# 4) 產生 Excel：含標題列、表頭樣式、千分位、紅字、總計列等
# 5) 設定會記錄到本機 JSON：管理員帳號、特殊店家設定、手動台數
# ============================================================

# --- 超級密碼(內部用) ---
# 1) 可查看總報表
# 2) 可繞過「管理員帳號=jjk888 禁用」的限制(避免外部客戶直接用後台帳)
SUPER_PASSWORD = "ccycs"

# --- 本機設定檔 ---
CONFIG_FILE = "config_settings.json"
DEBUG = True


# ============================================================
# 設定存取：記住使用者上次輸入
# ============================================================
def save_data():
    """將管理員帳號、特殊店家名單與手動台數存入 JSON"""
    cache = {
        "admin_acc": entry_acc.get().strip(),
        "special_configs": special_configs_data,
        "manual_terminals": entry_terminal.get().strip()
    }
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=4)
    update_special_count()


def load_data():
    """程式啟動時載入舊資料，避免每次重打"""
    if not os.path.exists(CONFIG_FILE):
        return

    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cache = json.load(f)

        # 1) 管理員帳號
        entry_acc.insert(0, cache.get("admin_acc", ""))

        # 2) 手動台數
        entry_terminal.delete(0, tk.END)
        entry_terminal.insert(0, cache.get("manual_terminals", ""))

        # 3) 特殊店家設定
        saved_configs = cache.get("special_configs", {})
        special_configs_data.update(saved_configs)

        for name, cfg in saved_configs.items():
            day_text = "月底" if cfg["day"] == 0 else "01號"
            special_listbox.insert(tk.END, f"{name} ({day_text} {cfg['hr']}點)")

    except:
        # 讀檔失敗就忽略，不影響主功能
        pass


# ============================================================
# 主流程：抓資料 -> 計算 -> 產 Excel
# ============================================================
def write_log(msg):
    log_text.config(state="normal")
    log_text.insert(tk.END, msg + "\n")
    log_text.see(tk.END)  # 自動捲到底
    log_text.config(state="disabled")

def run_combined_crawler(st_dt, ed_dt, admin_acc, status_label, btn, special_config, manual_terminals):
    """
    st_dt / ed_dt:
        UI 組合出的字串時間 "YYYY-mm-dd HH:MM:SS"
        st_dt 固定 01 號 + 08:00 (或 UI 設定)
        ed_dt 可選日期 + 08:00 (或 UI 設定)

    admin_acc:
        使用者輸入的帳號
        - 若輸入 ahp0369，實際抓取權限用 jjk888 (對外顯示帳號 vs 後台帳號)
        - 若輸入 jjk888，外部客戶不允許(避免猜到後台帳)，除非輸入 SUPER_PASSWORD

    special_config:
        特殊結算店家設定 dict
        例：{"某店": {"day": 0/1, "hr": 0-23}, ...}

    manual_terminals:
        手動輸入台數(選填)，若無則用店家台數加總
    """
    # --- API 設定 (token 寫死在程式內) ---
    CONFIG = {
        "banknote": {
            "url": "https://wpapi.ldjzmr.top/master/banknote_log",
            "token": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczovL3dwYXBpLmxkanptci50b3AvbWFzdGVyL2xvZ2luIiwiaWF0IjoxNzY4NjIxMzQzLCJleHAiOjE4MDAxNTczNDMsIm5iZiI6MTc2ODYyMTM0MywianRpIjoiaTNjNWdJcGp1M3Rsd0d4YyIsInN1YiI6IjEyIiwicHJ2IjoiMTg4ODk5NDM5MDUwZTVmMzc0MDliMThjYzZhNDk1NjkyMmE3YWIxYiJ9.qJaiec-CyY-yEtyur2SnbSpvqwaclT8huHGOsJinzjg"
        },
        "brand": {
            "url": "https://wpapi.ldjzmr.top/master/brand",
            "token": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJodHRwczovL3dwYXBpLmxkanptci50b3AvbWFzdGVyL2xvZ2luIiwiaWF0IjoxNzY4NjIxMzQzLCJleHAiOjE4MDAxNTczNDMsIm5iZiI6MTc2ODYyMTM0MywianRpIjoiaTNjNWdJcGp1M3Rsd0d4YyIsInN1YiI6IjEyIiwicHJ2IjoiMTg4ODk5NDM5MDUwZTVmMzc0MDliMThjYzZhNDk1NjkyMmE3YWIxYiJ9.qJaiec-CyY-yEtyur2SnbSpvqwaclT8huHGOsJinzjg"
        }
    }

    try:
        # UI：鎖按鈕、更新狀態
        btn.config(state="disabled")
        status_label.config(text="📡 正在連接 API...")

        raw_input_acc = admin_acc.strip()

        # --- 安全限制：禁止客戶直接用後台帳號 jjk888 ---
        # 注意：這個限制擋不住「修改程式/封包」的人，但能擋一般客戶誤用/亂輸
        if raw_input_acc == "jjk888" and raw_input_acc != SUPER_PASSWORD:
            messagebox.showerror("錯誤", "此帳號不可使用，請使用 a 開頭的新帳號。")
            return

        # --- 帳號映射：對外帳號 ahp0369 -> 實際權限帳號 jjk888 ---
        # 你的需求是：客戶可用 ahp0369 產表，但權限仍走 jjk888
        fetch_acc = "jjk888" if raw_input_acc == "ahp0369" else raw_input_acc

        # 解析結束時間
        dt_end = datetime.strptime(ed_dt, "%Y-%m-%d %H:%M:%S")

        # 判定：是否「新月份第一天結帳」
        # 用途：若是 1 號 08:00 的結帳點，前日累計就不做 (避免跨月干擾)
        is_new_month_start = (dt_end.day == 1 and dt_end.hour >= 8)

        # 前日切點：結束時間往前 24 小時(通常是 08:00)
        dt_offset_end = (dt_end - timedelta(hours=24)).strftime("%Y-%m-%d %H:%M:%S")

        # ========================================================
        # 1) 抓品牌(店家)資料：拿店名、管理員帳號、台數、創立時間
        # ========================================================
        brand_headers = {"Authorization": CONFIG["brand"]["token"], "User-Agent": "Mozilla/5.0"}
        brand_res = requests.get(
            CONFIG["brand"]["url"],
            headers=brand_headers,
            params={"page": 1, "page_size": 1000},
            timeout=20
        )
        brand_data = brand_res.json().get('data', {}).get('data', [])

        brand_mapping = []
        for b in brand_data:
            b_name = b.get('name')
            if not b_name:
                continue

            # 管理員帳號：你用 member.phone 來代表可見範圍
            # 台數：terminal_count
            brand_mapping.append({
                'name': b_name,
                '管理員帳號': str(b.get('member', {}).get('phone', '無')) if b.get('member') else '無',
                '台數': int(b.get('terminal_count', 0)),
                '創立時間': (b.get('created_at', '無') or '無')[:19]  # 取到秒
            })

        if not brand_mapping:
            raise ValueError("無法取得任何店家名稱資料")

        df_brand_map = pd.DataFrame(brand_mapping).drop_duplicates(subset=['name'])

        # ========================================================
        # 2) 抓帳務(banknote_log)資料：多頁 + 多執行緒
        # ========================================================
        status_label.config(text="📥 正在抓取帳單數據...")
        banknote_headers = {"Authorization": CONFIG["banknote"]["token"], "User-Agent": "Mozilla/5.0"}

        # 先抓第一頁拿總頁數 last_page
        init_res = requests.get(
            CONFIG["banknote"]["url"],
            headers=banknote_headers,
            params={"pagenum": 1, "pagesize": 100}
        )
        total_pages = init_res.json()['data']['list']['last_page']

        all_raw_banknote = []

        # stop_event：提早停止(如果抓到資料時間已早於 st_dt 很多)
        stop_event = threading.Event()

        def fetch_banknote_worker(page):
            """
            每個 worker 抓一頁 banknote_log
            如果資料已經早於 st_dt - 2days，觸發 stop_event 提早收工
            """
            if stop_event.is_set():
                return []

            try:
                r = requests.get(
                    CONFIG["banknote"]["url"],
                    headers=banknote_headers,
                    params={"pagenum": page, "pagesize": 500},
                    timeout=30
                )
                data = r.json().get('data', {}).get('list', {}).get('data', [])

                # 提早停止條件：最後一筆 created_at 已經比 st_dt 早很多
                if data:
                    cutoff = (datetime.strptime(st_dt, "%Y-%m-%d %H:%M:%S") - timedelta(days=2)).strftime("%Y-%m-%d %H:%M:%S")
                    if str(data[-1].get('created_at', '')) < cutoff:
                        stop_event.set()

                return data

            except:
                return []

        with ThreadPoolExecutor(max_workers=10) as executor:
            futures = [executor.submit(fetch_banknote_worker, p) for p in range(1, total_pages + 1)]
            for f in as_completed(futures):
                all_raw_banknote.extend(f.result())

        # ========================================================
        # 3) 彙總計算：依店家分組
        # ========================================================
        status_label.config(text="📊 正在計算財務數據...")

        full_df = pd.DataFrame(all_raw_banknote).drop_duplicates(subset=['id'])
        full_df['amount'] = pd.to_numeric(full_df['amount'], errors='coerce').fillna(0)

        def get_brand_name(x):
            return x.get('name', "未知") if isinstance(x, dict) else "未知"

        full_df['店家'] = full_df['brand'].apply(get_brand_name)

        # 先用 st_dt~ed_dt 篩一次，後面 groupby 再做特殊結算
        df_range_a = full_df[
            (full_df['created_at'].astype(str) >= st_dt) &
            (full_df['created_at'].astype(str) <= ed_dt)
        ].copy()


        if df_range_a.empty:
            messagebox.showwarning("提示", "此時間範圍內無任何交易數據")
            return

        spec_map = special_config if isinstance(special_config, dict) else {}
        # 預先宣告，避免 UnboundLocalError
        df_zero = pd.DataFrame(columns=[
            '店家', '開分', '投鈔', '洗分',
            '月初至今日累計營業額', '前日累計額', '今日變化'
        ])


        report_rows = []
        for brand, _group in df_range_a.groupby('店家'):
            # 預設起算時間 = UI 選的 st_dt (通常 01號 08:00)
            current_brand_st = st_dt

            # 若為特殊店：起算點可改為「月底/01號 + 指定小時」
            if brand in spec_map:
                cfg = spec_map[brand]
                st_obj = datetime.strptime(st_dt, "%Y-%m-%d %H:%M:%S")

                if cfg["day"] == 0:
                    # 月底：01號往前一天 = 上月最後一天
                    brand_st_obj = (st_obj - timedelta(days=1)).replace(hour=cfg["hr"], minute=0, second=0)
                else:
                    # 01號：同一天，只換小時
                    brand_st_obj = st_obj.replace(hour=cfg["hr"], minute=0, second=0)

                current_brand_st = brand_st_obj.strftime("%Y-%m-%d %H:%M:%S")

            # 3-1) 今日累計(當月累計)：從起算點 -> ed_dt
            brand_data = full_df[
                (full_df['店家'] == brand) &
                (full_df['created_at'].astype(str) >= current_brand_st) &
                (full_df['created_at'].astype(str) <= ed_dt)
            ]

            v_in = brand_data[brand_data['currency_type'] == 1]['amount'].sum()   # 投鈔
            v_open = brand_data[brand_data['currency_type'] == 2]['amount'].sum() # 開分
            v_wash = brand_data[brand_data['currency_type'] == 3]['amount'].sum() # 洗分
            accumulated = int(v_open - v_wash + v_in)

            # 3-2) 前日累計：從起算點 -> dt_offset_end
            prev_accum = 0
            if not is_new_month_start:
                brand_data_prev = full_df[
                    (full_df['店家'] == brand) &
                    (full_df['created_at'].astype(str) >= current_brand_st) &
                    (full_df['created_at'].astype(str) <= dt_offset_end)
                ]

                p_in = brand_data_prev[brand_data_prev['currency_type'] == 1]['amount'].sum()
                p_open = brand_data_prev[brand_data_prev['currency_type'] == 2]['amount'].sum()
                p_wash = brand_data_prev[brand_data_prev['currency_type'] == 3]['amount'].sum()
                prev_accum = int(p_open - p_wash + p_in)

            daily_change = accumulated - prev_accum

            report_rows.append({
                '店家': brand,
                '開分': int(v_open),
                '投鈔': int(v_in),
                '洗分': int(v_wash),
                '月初至今日累計營業額': accumulated,
                '前日累計額': prev_accum,
                '今日變化': daily_change
            })

        df_report = pd.DataFrame(report_rows)

        # 補上品牌資訊(管理員帳號/台數/創立時間)
        df_report = pd.merge(
            df_report,
            df_brand_map,
            left_on='店家',
            right_on='name',
            how='left'
        ).drop(columns=['name'])

        # 依創立時間排序(你用作下場順序的近似替代)
        df_report = df_report.sort_values(by='創立時間', ascending=True, na_position='last')

        # --- 權限過濾 ---
        # 非超級密碼：只顯示屬於該管理員帳號(fetch_acc)的店
        target_acc = admin_acc.strip()
        if raw_input_acc != SUPER_PASSWORD:
            df_report = df_report[df_report['管理員帳號'] == fetch_acc]

        if df_report.empty:
            messagebox.showwarning("提示", "權限範圍內無符合店家數據")
            return
        if raw_input_acc != SUPER_PASSWORD:
            df_report = df_report[df_report['管理員帳號'] == fetch_acc]

        if df_report.empty:
            messagebox.showwarning("提示", "權限範圍內無符合店家數據")
            return
        # ========================================================
        # 在這裡算 missing_names，建立 df_zero，然後 concat 回主表
        # ========================================================
        df_all = df_brand_map.copy()
        if raw_input_acc != SUPER_PASSWORD:
            df_all = df_all[df_all["管理員帳號"] == fetch_acc]

        all_names = set(df_all["name"].dropna().astype(str).tolist())
        shown_names = set(df_report["店家"].dropna().astype(str).tolist())

        df_zero = pd.DataFrame(columns=[
            '店家', '開分', '投鈔', '洗分',
            '月初至今日累計營業額', '前日累計額', '今日變化'
        ])
        
        # ========================================================
        # ✅ 補上未開分店家（帶創立時間），再整體依創立時間排序
        # ========================================================
        df_all = df_brand_map.copy()
        if raw_input_acc != SUPER_PASSWORD:
            df_all = df_all[df_all["管理員帳號"] == fetch_acc]

        # df_report 已經是權限過濾後的主表（且已有 創立時間/台數/管理員帳號）
        shown = set(df_report["店家"].dropna().astype(str).tolist())

        # 找出未出現的店家：直接用 df_all 來保證帶有「創立時間」
        df_missing = df_all[~df_all["name"].isin(shown)].copy()

        # 依 created_at 排序
        df_missing["創立時間"] = pd.to_datetime(df_missing["創立時間"], errors="coerce")
        df_missing = df_missing.sort_values(by="創立時間", ascending=True, na_position="last")

        # 轉成你主表要的欄位（把 name 改成 店家），其餘金額欄補 0
        df_zero = pd.DataFrame({
            "店家": df_missing["name"].astype(str),
            "開分": 0,
            "投鈔": 0,
            "洗分": 0,
            "月初至今日累計營業額": 0,
            "前日累計額": 0,
            "今日變化": 0,
            "管理員帳號": df_missing["管理員帳號"].astype(str),
            "台數": pd.to_numeric(df_missing["台數"], errors="coerce").fillna(0).astype(int),
            "創立時間": df_missing["創立時間"]
        })

        # 合併回主表
        df_report = pd.concat([df_report, df_zero], ignore_index=True)

        # ✅ 合併後再整體排序，0 店家才會插回正確位置
        df_report["創立時間"] = pd.to_datetime(df_report["創立時間"], errors="coerce")
        df_report = df_report.sort_values(by="創立時間", ascending=True, na_position="last").reset_index(drop=True)

        # missing_names 給右側清單用（順序就是 created_at）
        missing_names = df_zero["店家"].tolist()



        # --- 加總列 ---
        summary = {
            '店家': '總計',
            '開分': df_report['開分'].sum(),
            '投鈔': df_report['投鈔'].sum(),
            '洗分': df_report['洗分'].sum(),
            '月初至今日累計營業額': df_report['月初至今日累計營業額'].sum(),
            '前日累計額': df_report['前日累計額'].sum(),
            '今日變化': df_report['今日變化'].sum(),
            '創立時間': ''
        }
        df_final = pd.concat([df_report, pd.DataFrame([summary])], ignore_index=True)

        # 台數：可手動輸入覆蓋
        total_shops = len(df_report)
        if manual_terminals and manual_terminals.isdigit():
            total_terminals = int(manual_terminals)
        else:
            total_terminals = df_report['台數'].sum()

        nickname = "總報表" if target_acc == SUPER_PASSWORD else target_acc

        # ========================================================
        # 4) 輸出 Excel：營業狀況表
        # ========================================================
        file_date = dt_end.strftime("%-m.%-d") if os.name != "nt" else dt_end.strftime("%#m.%#d")
        file_time = dt_end.strftime("%H-%M")

        out_file = f"{file_date} 開洗分狀況({file_time}).xlsx"



        with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
            # 寫資料(從第4列下方開始)，讓上面留空做標題/說明
            display_cols = ['店家', '開分', '投鈔', '洗分', '月初至今日累計營業額', '前日累計額', '今日變化']
            df_final[display_cols].to_excel(writer, sheet_name="營業狀況", index=False, startrow=3)
            ws = writer.sheets["營業狀況"]
        # =========================
        # 右側：未開分店家清單
        # =========================
        # 1) 該帳號名下「應有店家」
            df_all = df_brand_map.copy()

            # 注意：你的 df_brand_map 管理員帳號欄是字串 phone
            # fetch_acc 是你用來比對權限的帳號（ahp0369 會映射 jjk888）
            if raw_input_acc != SUPER_PASSWORD:
                df_all = df_all[df_all["管理員帳號"] == fetch_acc]

            all_names = set(df_all["name"].dropna().astype(str).tolist())

            # 2) 主表「已出現店家」
            shown_names = set(df_report["店家"].dropna().astype(str).tolist())

            # 4) 寫到 Excel 右側（例如 I4 開始）
            start_col = 9  # I
            start_row = 4

            title_cell = ws.cell(row=start_row, column=start_col)
            title_cell.value = "未開分店家清單"
            title_cell.font = Font(name="微軟正黑體", bold=True, size=12)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 2)

            # 內容往下列
            for i, name in enumerate(missing_names, start=1):
                c = ws.cell(row=start_row + i, column=start_col)
                c.value = name
                c.font = Font(name="微軟正黑體", size=11)
                c.alignment = Alignment(horizontal="left", vertical="center")

            # 欄寬調整
            ws.column_dimensions["I"].width = 18
            ws.column_dimensions["J"].width = 2
            ws.column_dimensions["K"].width = 2
            ws.cell(row=start_row+1, column=start_col+1).value = f"共 {len(missing_names)} 家"


            # ---------- Excel 全域字體 ----------
            ms_font = Font(name='微軟正黑體', size=12)
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = ms_font

            # ---------- Page setup ----------
            ws.page_margins.left = 0.1
            ws.page_margins.right = 0.1
            ws.page_margins.top = 0.25
            ws.page_margins.bottom = 0.25
            ws.page_margins.header = 0.1
            ws.page_margins.footer = 0.1
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_setup.horizontalCentered = True

            # ---------- 頂部標題 ----------
            now_str = dt_end.strftime("%Y/%m/%d %H:%M")
            ws.merge_cells('A1:B1')
            ws['A1'] = now_str
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells('C1:E1')
            ws['C1'] = "營業狀況(當月累計)"
            ws['C1'].alignment = Alignment(horizontal='center')

            info_font = Font(name='微軟正黑體', size=14)
            ws['F1'] = f"已下場台數: {total_terminals}"
            ws['F1'].font = info_font
            ws['F2'] = f"店家數: {total_shops}"
            ws['F2'].font = info_font

            ws.merge_cells('C2:E2')
            ws['C2'] = "PS:有含開洗分測試"
            ws['C2'].font = Font(name='微軟正黑體', color="FF0000", size=14)
            ws['C2'].alignment = Alignment(horizontal='center')

            ws['A1'].font = Font(name='微軟正黑體', size=14)
            ws['C1'].font = Font(name='微軟正黑體', size=14)
            ws['C2'].font = Font(name='微軟正黑體', size=14, color="FF0000")

            # ---------- 樣式 ----------
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            shop_col_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

            headers = ["店家", "開分", "投鈔", "洗分", "月初至今日\n累計營業額", "前日累計額", "今日變化"]

            # 表頭：合併第3~4列
            for col_idx, header_text in enumerate(headers, 1):
                col_letter = ws.cell(row=3, column=col_idx).column_letter
                cell_3 = ws.cell(row=3, column=col_idx)
                cell_4 = ws.cell(row=4, column=col_idx)

                cell_3.value = header_text
                ws.merge_cells(f"{col_letter}3:{col_letter}4")

                # 合併時要讓上下半都畫到邊框，避免黑線斷
                cell_3.border = border
                cell_4.border = border

                cell_3.fill = header_fill
                cell_3.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                cell_3.font = Font(color="000000", size=12, name='微軟正黑體')

            # 欄寬(依你客戶需求固定)
            for col_idx in range(1, 8):
                col_letter = ws.cell(row=3, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = 11.01 if col_idx == 1 else 15.72

            # 行高
            ws.row_dimensions[1].height = 19.75
            ws.row_dimensions[2].height = 19.75
            ws.row_dimensions[3].height = 19.75
            ws.row_dimensions[4].height = 19.75
            for r in range(5, ws.max_row + 1):
                ws.row_dimensions[r].height = 19.5
            ws.print_title_rows = '3:4'

            # ---------- 資料區樣式 ----------
            for col_idx in range(1, 8):
                for row_idx in range(4, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = border

                    # 對齊：店家置中、數字靠右
                    if col_idx == 1:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='right', vertical='center')

                    # 店家欄背景
                    if col_idx == 1 and row_idx > 4:
                        cell.fill = shop_col_fill

                    # 總計列背景(最後一列)
                    if row_idx == ws.max_row:
                        total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
                        cell.fill = total_fill
                        cell.font = Font(color="000000", name='微軟正黑體')

                    # 數字格式 + 紅字
                    if row_idx > 4 and col_idx >= 2:
                        val = cell.value
                        if val is None:
                            val = 0

                        if isinstance(val, (int, float)):
                            # 今日變化欄：正數要帶 + ，0 也要顯示 +0
                            if col_idx == 7:
                                cell.number_format = '"+ "#,##0;[Red]-#,##0;"+ "0'
                            # 投鈔欄：0 顯示空白
                            elif col_idx == 3:
                                cell.number_format = '#,##0;-#,##0;""'
                            else:
                                cell.number_format = '#,##0'

                            is_last_row = (row_idx == ws.max_row)
                            font_size = 12

                            if val < 0:
                                cell.font = Font(color="FF0000", size=font_size, name='微軟正黑體')
                            else:
                                cell.font = Font(color="000000", size=font_size, name='微軟正黑體')

        status_label.config(text="✅ 成功生成")
        messagebox.showinfo("成功", f"報表已產出：{out_file}")

    except Exception as e:
        messagebox.showerror("錯誤", f"發生問題: {str(e)}")

    finally:
        btn.config(state="normal")
        status_label.config(text="就緒")


# ============================================================
# UI：按鈕觸發主流程
# ============================================================
def start_process():
    """按下按鈕後：先存設定，再開 thread 跑主流程(避免 UI 卡住)"""
    save_data()

    acc = entry_acc.get().strip()
    if not acc:
        messagebox.showerror("錯誤", "請輸入帳號")
        return

    st = f"{cb_year.get()}-{cb_month.get()}-01 {h_st.get()}:{m_st.get()}:{s_st.get()}"
    ed = f"{cal_ed.get_date()} {h_ed.get()}:{m_ed.get()}:{s_ed.get()}"
    manual_terminals = entry_terminal.get().strip()

    threading.Thread(
        target=run_combined_crawler,
        args=(st, ed, acc, status_label, btn, special_configs_data, manual_terminals),
        daemon=True
    ).start()
def update_special_count():
    count = len(special_configs_data)
    f_special.config(text=f" 🏪 特殊結算店家清單（共 {count} 家）")

# ============================================================
# UI：主視窗
# ============================================================
root = tk.Tk()
root.title("王牌財務工具 V2.1")
root.geometry("450x650")

# --- 管理員帳號 ---
f_acc = tk.LabelFrame(root, text=" 管理員帳號", padx=10, pady=10)
f_acc.pack(pady=10, padx=20, fill="x")
entry_acc = tk.Entry(f_acc, font=("Arial", 12))
entry_acc.pack(fill="x")

# --- 手動台數 ---
f_terminal = tk.LabelFrame(root, text=" 🎰 手動設定台數 (選填)", padx=10, pady=10)
f_terminal.pack(pady=5, padx=20, fill="x")
entry_terminal = tk.Entry(f_terminal, font=("Arial", 12))
entry_terminal.pack(fill="x")
tk.Label(f_terminal, text="", fg="gray", font=("微軟正黑體", 8)).pack()

# --- 開始時間(固定 01 號) ---
f_st = tk.LabelFrame(root, text=" 開始時間 (日期固定為 01 號)", padx=10, pady=10)
f_st.pack(pady=10, padx=20, fill="x")

this_year = datetime.now().year
cb_year = ttk.Combobox(f_st, values=[this_year - 1, this_year, this_year + 1], width=7, state="readonly")
cb_year.set(this_year)
cb_year.pack(side="left", padx=2)

cb_month = ttk.Combobox(f_st, values=[f"{i:02d}" for i in range(1, 13)], width=4, state="readonly")
cb_month.set(f"{datetime.now().month:02d}")
cb_month.pack(side="left", padx=2)

h_st = ttk.Spinbox(f_st, from_=0, to=23, width=3, format="%02.0f")
h_st.set("08")
h_st.pack(side="left", padx=2)

m_st = ttk.Spinbox(f_st, from_=0, to=59, width=3, format="%02.0f")
m_st.set("00")
m_st.pack(side="left", padx=2)

s_st = ttk.Spinbox(f_st, from_=0, to=59, width=3, format="%02.0f")
s_st.set("00")
s_st.pack(side="left", padx=2)

# --- 結束時間 ---
f_ed = tk.LabelFrame(root, text=" 結束時間", padx=10, pady=10)
f_ed.pack(pady=10, padx=20, fill="x")

cal_ed = DateEntry(f_ed, width=12, date_pattern='yyyy-mm-dd')
cal_ed.pack(side="left", padx=2)

h_ed = ttk.Spinbox(f_ed, from_=0, to=23, width=3, format="%02.0f")
h_ed.set("08")
h_ed.pack(side="left", padx=2)

m_ed = ttk.Spinbox(f_ed, from_=0, to=59, width=3, format="%02.0f")
m_ed.set("00")
m_ed.pack(side="left", padx=2)

s_ed = ttk.Spinbox(f_ed, from_=0, to=59, width=3, format="%02.0f")
s_ed.set("00")
s_ed.pack(side="left", padx=2)

# ============================================================
# UI：特殊店家設定(可新增/刪除/雙擊編輯)
# ============================================================
special_title_var = tk.StringVar()
special_title_var.set(" 🏪 特殊結算店家清單（共 0 家）")

f_special = tk.LabelFrame(root, text=" 🏪 特殊結算店家清單（共 0 家）", padx=10, pady=10)
f_special.pack(pady=5, padx=20, fill="x")

list_frame = tk.Frame(f_special)
list_frame.pack(fill="x")

special_listbox = tk.Listbox(list_frame, height=4, font=("微軟正黑體", 10))
special_listbox.pack(side="left", fill="x", expand=True)

scrollbar = tk.Scrollbar(list_frame)
scrollbar.pack(side="right", fill="y")
special_listbox.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=special_listbox.yview)

btn_frame = tk.Frame(f_special)
btn_frame.pack(fill="x", pady=5)

# 全域字典：記錄特殊店家設定
special_configs_data = {}


def update_listbox_display():
    """根據 special_configs_data 重新整理 Listbox"""
    special_listbox.delete(0, tk.END)
    for name, cfg in special_configs_data.items():
        day_text = "月底" if cfg["day"] == 0 else "01號"
        special_listbox.insert(tk.END, f"{name} ({day_text} {cfg['hr']}點)")


def add_special_shop(edit_name=None, edit_data=None):
    """新增/編輯特殊店家結算設定"""
    win = tk.Toplevel(root)
    win.title("編輯店家設定" if edit_name else "新增特殊店家")
    win.geometry("300x250")
    win.grab_set()

    tk.Label(win, text="店家名稱:").pack(pady=5)
    name_ent = tk.Entry(win)
    if edit_name:
        name_ent.insert(0, edit_name)
    name_ent.pack()

    tk.Label(win, text="結算日期:").pack(pady=5)
    initial_day = str(edit_data["day"]) if edit_data else "1"
    day_var = tk.StringVar(value=initial_day)

    tk.Radiobutton(win, text="月底結帳", variable=day_var, value="0").pack()
    tk.Radiobutton(win, text="01 號結帳", variable=day_var, value="1").pack()

    tk.Label(win, text="結算小時 (0-23):").pack(pady=5)
    hr_spin = ttk.Spinbox(win, from_=0, to=23, width=5)
    hr_spin.set(edit_data["hr"] if edit_data else 0)
    hr_spin.pack()

    def save():
        name = name_ent.get().strip()
        if not name:
            return

        # 編輯模式：先刪舊 key 再寫新 key
        if edit_name and edit_name in special_configs_data:
            del special_configs_data[edit_name]

        special_configs_data[name] = {"day": int(day_var.get()), "hr": int(hr_spin.get())}

        update_listbox_display()
        save_data()
        win.destroy()

    tk.Button(win, text="儲存設定", command=save, bg="#4CAF50", fg="white").pack(pady=10)


def clear_special_shop():
    """清空所有特殊店家設定"""
    special_listbox.delete(0, tk.END)
    special_configs_data.clear()
    save_data()
    update_special_count()


def delete_selected_shop():
    """刪除 Listbox 目前選中的店家設定"""
    selection = special_listbox.curselection()
    if not selection:
        messagebox.showwarning("提示", "請先點選清單中要刪除的店家")
        return
    update_special_count()

    index = selection[0]
    item_text = special_listbox.get(index)
    shop_name = item_text.split(" (")[0]

    if shop_name in special_configs_data:
        del special_configs_data[shop_name]

    special_listbox.delete(index)
    save_data()


tk.Button(btn_frame, text="＋ 新增店家", command=add_special_shop, bg="#E1F5FE").pack(side="left", padx=5)
tk.Button(btn_frame, text="➖ 刪除選中", command=delete_selected_shop).pack(side="left", padx=5)
tk.Button(btn_frame, text="🗑️ 清除全部", command=clear_special_shop).pack(side="left")


# ============================================================
# UI：產出按鈕 + 狀態列
# ============================================================
btn = tk.Button(
    root,
    text=" 生成對帳報表",
    command=start_process,
    bg="#2196F3",
    fg="white",
    font=("Arial", 12, "bold"),
    height=2
)
btn.pack(pady=20, padx=20, fill="x")

status_label = tk.Label(root, text="就緒", fg="gray")
status_label.pack()
# ============================================================
# UI：日誌區（Debug / 流程顯示）
# ============================================================
f_log = tk.LabelFrame(root, text=" 執行日誌 ", padx=5, pady=5)
f_log.pack(padx=10, pady=5, fill="both", expand=True)

log_text = tk.Text(
    f_log,
    height=8,
    font=("Consolas", 9),
    state="disabled",
    wrap="word"
)
log_text.pack(fill="both", expand=True)


# ============================================================
# Listbox 雙擊：直接進入編輯
# ============================================================
def on_double_click(_event):
    selection = special_listbox.curselection()
    if selection:
        index = selection[0]
        item_text = special_listbox.get(index)
        shop_name = item_text.split(" (")[0]
        shop_data = special_configs_data.get(shop_name)
        add_special_shop(edit_name=shop_name, edit_data=shop_data)


special_listbox.bind('<Double-1>', on_double_click)

# 啟動時載入設定
load_data()
update_special_count()

root.mainloop()
